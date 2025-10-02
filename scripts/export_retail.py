#!/usr/bin/env python3
"""
POC Retail BI - Export Script
==============================
Extraction des données retail depuis BDD ou CSV vers Excel
Génère un rapport automatisé avec KPIs et détails des ventes

Author: Florent (Data Analyst POC)
Date: October 2024
"""

import pandas as pd
import psycopg2
import sqlite3
from sqlalchemy import create_engine
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, LineChart, Reference
from datetime import datetime
import os
import sys
from pathlib import Path
from dotenv import load_dotenv

# Chargement des variables d'environnement
load_dotenv()

# Configuration sécurisée
DB_CONFIG = {
    'host': os.getenv('DB_HOST', 'localhost'),
    'database': os.getenv('DB_DATABASE', 'retail_demo'),
    'user': os.getenv('DB_USER', 'postgres'),
    'password': os.getenv('DB_PASSWORD', 'password'),
    'port': int(os.getenv('DB_PORT', 5432))
}

# Chemins des fichiers
CURRENT_DIR = Path(__file__).parent
PROJECT_ROOT = CURRENT_DIR.parent
DATA_DIR = PROJECT_ROOT / 'data'
OUTPUT_DIR = PROJECT_ROOT / 'outputs'
CSV_FILE = DATA_DIR / 'retail_demo.csv'
EXCEL_OUTPUT = OUTPUT_DIR / 'reporting_retail.xlsx'

def connect_to_database():
    """
    Tentative de connexion à la base PostgreSQL
    Retourne None si échec (fallback vers CSV)
    """
    try:
        engine = create_engine(
            f"postgresql://{DB_CONFIG['user']}:{DB_CONFIG['password']}"
            f"@{DB_CONFIG['host']}:{DB_CONFIG['port']}/{DB_CONFIG['database']}"
        )
        # Test de connexion
        with engine.connect() as conn:
            conn.execute("SELECT 1")
        print("✅ Connexion PostgreSQL réussie")
        return engine
    except Exception as e:
        print(f"❌ Erreur connexion PostgreSQL: {e}")
        print("🔄 Basculement vers fichier CSV...")
        return None

def load_data_from_database(engine):
    """
    Charge les données depuis la base PostgreSQL
    Utilise les vues créées dans le script SQL
    """
    try:
        # Requête principale: ventes complètes avec jointures
        query_ventes = """
        SELECT 
            v.id_vente,
            v.date_vente,
            EXTRACT(YEAR FROM v.date_vente) as annee,
            EXTRACT(MONTH FROM v.date_vente) as mois,
            TO_CHAR(v.date_vente, 'YYYY-MM') as periode,
            c.nom || ' ' || c.prenom as nom_client,
            c.email,
            c.ville as ville_client,
            c.age,
            e.nom_enseigne,
            e.ville as ville_enseigne,
            e.region,
            v.produit,
            v.quantite,
            v.prix_unitaire,
            v.montant_total
        FROM vente v
        JOIN client c ON v.id_client = c.id_client
        JOIN enseigne e ON v.id_enseigne = e.id_enseigne
        ORDER BY v.date_vente DESC
        """
        
        df_ventes = pd.read_sql(query_ventes, engine)
        
        # KPIs par enseigne
        query_kpis = """
        SELECT 
            e.nom_enseigne,
            e.ville,
            e.region,
            COUNT(DISTINCT v.id_client) as nb_clients,
            COUNT(v.id_vente) as nb_ventes,
            COALESCE(SUM(v.montant_total), 0) as chiffre_affaires,
            COALESCE(AVG(v.montant_total), 0) as panier_moyen
        FROM enseigne e
        LEFT JOIN vente v ON e.id_enseigne = v.id_enseigne
        GROUP BY e.id_enseigne, e.nom_enseigne, e.ville, e.region
        ORDER BY chiffre_affaires DESC
        """
        
        df_kpis = pd.read_sql(query_kpis, engine)
        
        # Évolution mensuelle
        query_evolution = """
        SELECT 
            EXTRACT(YEAR FROM v.date_vente) as annee,
            EXTRACT(MONTH FROM v.date_vente) as mois,
            TO_CHAR(v.date_vente, 'YYYY-MM') as periode,
            COUNT(v.id_vente) as nb_ventes,
            SUM(v.montant_total) as chiffre_affaires,
            AVG(v.montant_total) as panier_moyen,
            COUNT(DISTINCT v.id_client) as nb_clients_uniques
        FROM vente v
        GROUP BY 
            EXTRACT(YEAR FROM v.date_vente),
            EXTRACT(MONTH FROM v.date_vente),
            TO_CHAR(v.date_vente, 'YYYY-MM')
        ORDER BY annee, mois
        """
        
        df_evolution = pd.read_sql(query_evolution, engine)
        
        print(f"✅ Données chargées depuis PostgreSQL:")
        print(f"   📊 {len(df_ventes)} ventes")
        print(f"   🏪 {len(df_kpis)} enseignes")
        print(f"   📅 {len(df_evolution)} périodes")
        
        return df_ventes, df_kpis, df_evolution
        
    except Exception as e:
        print(f"❌ Erreur lors du chargement depuis BDD: {e}")
        raise

def load_data_from_csv():
    """
    Charge les données depuis le fichier CSV de fallback
    Simule les mêmes calculs que la base de données
    """
    try:
        # Lecture du CSV principal
        df_ventes = pd.read_csv(CSV_FILE)
        
        # Conversion des dates
        df_ventes['date_vente'] = pd.to_datetime(df_ventes['date_vente'])
        df_ventes['annee'] = df_ventes['date_vente'].dt.year
        df_ventes['mois'] = df_ventes['date_vente'].dt.month
        df_ventes['periode'] = df_ventes['date_vente'].dt.strftime('%Y-%m')
        
        # Calcul des KPIs par enseigne
        df_kpis = df_ventes.groupby(['nom_enseigne', 'ville_enseigne', 'region']).agg({
            'nom_client': 'nunique',
            'id_vente': 'count',
            'montant_total': ['sum', 'mean']
        }).round(2)
        
        df_kpis.columns = ['nb_clients', 'nb_ventes', 'chiffre_affaires', 'panier_moyen']
        df_kpis = df_kpis.reset_index()
        df_kpis.columns = ['nom_enseigne', 'ville', 'region', 'nb_clients', 'nb_ventes', 'chiffre_affaires', 'panier_moyen']
        df_kpis = df_kpis.sort_values('chiffre_affaires', ascending=False)
        
        # Évolution mensuelle
        df_evolution = df_ventes.groupby(['annee', 'mois', 'periode']).agg({
            'id_vente': 'count',
            'montant_total': ['sum', 'mean'],
            'nom_client': 'nunique'
        }).round(2)
        
        df_evolution.columns = ['nb_ventes', 'chiffre_affaires', 'panier_moyen', 'nb_clients_uniques']
        df_evolution = df_evolution.reset_index()
        df_evolution = df_evolution.sort_values(['annee', 'mois'])
        
        print(f"✅ Données chargées depuis CSV:")
        print(f"   📊 {len(df_ventes)} ventes")
        print(f"   🏪 {len(df_kpis)} enseignes")
        print(f"   📅 {len(df_evolution)} périodes")
        
        return df_ventes, df_kpis, df_evolution
        
    except Exception as e:
        print(f"❌ Erreur lors du chargement depuis CSV: {e}")
        raise

def calculate_global_kpis(df_ventes, df_kpis):
    """
    Calcule les KPIs globaux pour le dashboard
    """
    kpis_globaux = {
        'ca_total': df_ventes['montant_total'].sum(),
        'nb_ventes_total': len(df_ventes),
        'nb_clients_total': df_ventes['nom_client'].nunique(),
        'nb_enseignes': len(df_kpis),
        'panier_moyen_global': df_ventes['montant_total'].mean(),
        'premiere_vente': df_ventes['date_vente'].min(),
        'derniere_vente': df_ventes['date_vente'].max(),
        'enseigne_top': df_kpis.iloc[0]['nom_enseigne'] if len(df_kpis) > 0 else 'N/A',
        'ca_top_enseigne': df_kpis.iloc[0]['chiffre_affaires'] if len(df_kpis) > 0 else 0
    }
    
    return kpis_globaux

def style_excel_sheet(ws, title, is_summary=False):
    """
    Applique un style professionnel aux feuilles Excel
    """
    # Style d'en-tête
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    
    # Style titre principal
    title_font = Font(bold=True, size=16, color="2F5597")
    
    if is_summary:
        # Style spécial pour la page de synthèse
        ws['A1'] = title
        ws['A1'].font = title_font
        ws.merge_cells('A1:D1')
        
        # Headers pour KPIs
        ws['A3'] = "Indicateur"
        ws['B3'] = "Valeur"
        
        for cell in ['A3', 'B3']:
            ws[cell].font = header_font
            ws[cell].fill = header_fill
            ws[cell].alignment = Alignment(horizontal='center')
    else:
        # Style pour les feuilles de données
        if ws.max_row > 0:
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
    
    # Auto-ajustement des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = None
        for cell in column:
            if hasattr(cell, 'column_letter'):
                column_letter = cell.column_letter
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        if column_letter:
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

def create_excel_report(df_ventes, df_kpis, df_evolution, kpis_globaux):
    """
    Génère le rapport Excel avec plusieurs onglets et mise en forme
    """
    # Création du fichier Excel
    OUTPUT_DIR.mkdir(exist_ok=True)
    
    with pd.ExcelWriter(EXCEL_OUTPUT, engine='openpyxl') as writer:
        
        # Onglet 1: Synthèse KPIs
        df_synthese = pd.DataFrame([
            ["Chiffre d'affaires total", f"{kpis_globaux['ca_total']:,.2f} €"],
            ["Nombre de ventes", f"{kpis_globaux['nb_ventes_total']:,}"],
            ["Nombre de clients", f"{kpis_globaux['nb_clients_total']:,}"],
            ["Nombre d'enseignes", f"{kpis_globaux['nb_enseignes']}"],
            ["Panier moyen", f"{kpis_globaux['panier_moyen_global']:,.2f} €"],
            ["Période d'analyse", f"{kpis_globaux['premiere_vente'].strftime('%d/%m/%Y')} - {kpis_globaux['derniere_vente'].strftime('%d/%m/%Y')}"],
            ["Top enseigne", f"{kpis_globaux['enseigne_top']}"],
            ["CA top enseigne", f"{kpis_globaux['ca_top_enseigne']:,.2f} €"],
        ], columns=['Indicateur', 'Valeur'])
        
        df_synthese.to_excel(writer, sheet_name='📊 Synthèse KPIs', index=False, startrow=2)
        
        # Onglet 2: KPIs par enseigne
        df_kpis_formatted = df_kpis.copy()
        df_kpis_formatted['chiffre_affaires'] = df_kpis_formatted['chiffre_affaires'].apply(lambda x: f"{x:,.2f} €")
        df_kpis_formatted['panier_moyen'] = df_kpis_formatted['panier_moyen'].apply(lambda x: f"{x:.2f} €")
        df_kpis_formatted.to_excel(writer, sheet_name='🏪 KPIs Enseignes', index=False)
        
        # Onglet 3: Évolution mensuelle
        df_evolution_formatted = df_evolution.copy()
        df_evolution_formatted['chiffre_affaires'] = df_evolution_formatted['chiffre_affaires'].apply(lambda x: f"{x:,.2f} €")
        df_evolution_formatted['panier_moyen'] = df_evolution_formatted['panier_moyen'].apply(lambda x: f"{x:.2f} €")
        df_evolution_formatted.to_excel(writer, sheet_name='📈 Évolution Mensuelle', index=False)
        
        # Onglet 4: Détail des ventes (top 100 pour lisibilité)
        df_ventes_top = df_ventes.head(100).copy()
        df_ventes_top['montant_total'] = df_ventes_top['montant_total'].apply(lambda x: f"{x:.2f} €")
        df_ventes_top['prix_unitaire'] = df_ventes_top['prix_unitaire'].apply(lambda x: f"{x:.2f} €")
        df_ventes_top.to_excel(writer, sheet_name='🛒 Détail Ventes (Top 100)', index=False)
        
        # Application des styles
        workbook = writer.book
        
        # Style onglet synthèse
        ws_synthese = workbook['📊 Synthèse KPIs']
        style_excel_sheet(ws_synthese, "🎯 DASHBOARD RETAIL BI - SYNTHÈSE EXÉCUTIVE", True)
        
        # Style autres onglets
        for sheet_name in ['🏪 KPIs Enseignes', '📈 Évolution Mensuelle', '🛒 Détail Ventes (Top 100)']:
            ws = workbook[sheet_name]
            style_excel_sheet(ws, sheet_name)
    
    print(f"✅ Rapport Excel généré: {EXCEL_OUTPUT}")
    print(f"📋 Onglets créés:")
    print(f"   🎯 Synthèse KPIs (vue exécutive)")
    print(f"   🏪 KPIs par Enseigne")
    print(f"   📈 Évolution Mensuelle")
    print(f"   🛒 Détail des Ventes")

def main():
    """
    Fonction principale - Orchestration du processus ETL
    """
    print("🚀 POC RETAIL BI - Export vers Excel")
    print("=" * 50)
    
    # Étape 1: Tentative de connexion BDD
    engine = connect_to_database()
    
    # Étape 2: Chargement des données
    if engine:
        df_ventes, df_kpis, df_evolution = load_data_from_database(engine)
        source = "PostgreSQL"
    else:
        df_ventes, df_kpis, df_evolution = load_data_from_csv()
        source = "CSV"
    
    # Étape 3: Calcul des KPIs globaux
    kpis_globaux = calculate_global_kpis(df_ventes, df_kpis)
    
    # Étape 4: Affichage résumé
    print(f"\n📈 RÉSUMÉ DE L'ANALYSE ({source}):")
    print("-" * 40)
    print(f"💰 CA Total: {kpis_globaux['ca_total']:,.2f} €")
    print(f"🛒 Ventes: {kpis_globaux['nb_ventes_total']:,}")
    print(f"👥 Clients: {kpis_globaux['nb_clients_total']:,}")
    print(f"🏪 Enseignes: {kpis_globaux['nb_enseignes']}")
    print(f"🎯 Panier moyen: {kpis_globaux['panier_moyen_global']:,.2f} €")
    print(f"🏆 Top enseigne: {kpis_globaux['enseigne_top']}")
    
    # Étape 5: Génération du rapport Excel
    print(f"\n📊 GÉNÉRATION DU RAPPORT EXCEL:")
    print("-" * 40)
    create_excel_report(df_ventes, df_kpis, df_evolution, kpis_globaux)
    
    # Étape 6: Instructions suivantes
    print(f"\n🎯 PROCHAINES ÉTAPES:")
    print("-" * 40)
    print(f"1. 📂 Ouvrir le fichier: {EXCEL_OUTPUT}")
    print(f"2. 🔌 Connecter Power BI au fichier Excel")
    print(f"3. 🌐 Publier le dashboard Power BI")
    print(f"4. 📤 Pusher le code sur GitHub")
    
    print(f"\n✨ Export terminé avec succès ! Source: {source}")

if __name__ == "__main__":
    main()